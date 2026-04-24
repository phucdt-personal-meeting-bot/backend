import io

from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock


def extract_text_cells(file_bytes: bytes) -> dict[str, list[dict]]:
    """Extract text cells from each sheet. Returns {sheet_name: [{ref, text}, ...]}."""
    wb = load_workbook(io.BytesIO(file_bytes), data_only=False)
    result = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        cells = []
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                if isinstance(cell.value, str) and not cell.value.startswith("="):
                    cells.append({"ref": cell.coordinate, "text": cell.value})
        result[sheet_name] = cells

    wb.close()
    return result


def _rebuild_rich_text(original: CellRichText, translated: str):
    """Reapply rich text formatting from original to translated text via substring matching."""
    formatted_runs = []
    for part in original:
        if isinstance(part, TextBlock) and part.font and part.text.strip():
            formatted_runs.append((part.text, part.font))

    if not formatted_runs:
        return translated

    regions = []
    for text, font in formatted_runs:
        idx = translated.find(text)
        if idx == -1:
            continue
        end = idx + len(text)
        if any(s < end and idx < e for s, e, _ in regions):
            continue
        regions.append((idx, end, font))

    if not regions:
        return translated

    regions.sort()

    parts = []
    pos = 0
    for start, end, font in regions:
        if start > pos:
            parts.append(translated[pos:start])
        parts.append(TextBlock(font, translated[start:end]))
        pos = end
    if pos < len(translated):
        parts.append(translated[pos:])

    return CellRichText(parts)


def write_translations(file_bytes: bytes, translations: dict[str, dict[str, str]]) -> bytes:
    """Write translated text back to the workbook, preserving rich text formatting."""
    wb_fmt = load_workbook(io.BytesIO(file_bytes), rich_text=True)
    wb = load_workbook(io.BytesIO(file_bytes))

    for sheet_name, cell_map in translations.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        ws_fmt = wb_fmt[sheet_name]
        for ref, text in cell_map.items():
            original = ws_fmt[ref].value
            if isinstance(original, CellRichText):
                ws[ref] = _rebuild_rich_text(original, text)
            else:
                ws[ref] = text

    output = io.BytesIO()
    wb.save(output)
    wb.close()
    wb_fmt.close()
    return output.getvalue()
